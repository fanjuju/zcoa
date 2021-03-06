import openpyxl

class ReadExcel:
    def get_excel(self):
        # 获取工作簿
        wb = openpyxl.load_workbook('../data/data.xlsx')
        # 获取sheet页
        ws = wb['新增供应商申请测试用例']
        # 列表套列表
        all_cases = []
        for i in ws.iter_rows(min_row=2,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
            row_list = [cell.value for cell in i]
            all_cases.append(row_list)
        return all_cases

if __name__ == '__main__':
    all_cases = ReadExcel().get_excel()
    print(all_cases)

